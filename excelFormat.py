import openpyxl as excel
import os
from smtplib import SMTP
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import time
#---------------------------------- D E C L A R A C I O N E S ----------------------------------------
# Para el servidor de correo
port = 25
smtp_server="pmi.org.ar"

#user = input("Enter user:")
user = "rosario@pmi.org.ar"
#password = input("Enter password:")
password = "rosario2323"

#message = MIMEMultipart("alternative")
#message["Subject"] = "Test mail personalizado"
#message["From"] = sender_email
#message["To"] = receiver_email

# Para el HTML del mail (para hacerlo mas lindo)
# Separar el HEAD y el BODY para poder editar tranquilamente el BODY con la funcion format()
HEAD="""\
<!DOCTYPE html>
<html xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml"
    xmlns:o="urn:schemas-microsoft-com:office:office">

<head>
    <title></title>
    <!--[if !mso]><!-- -->
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <!--<![endif]-->
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style type="text/css">
        #outlook a {
            padding: 0;
        }

        .ReadMsgBody {
            width: 100%;
        }

        .ExternalClass {
            width: 100%;
        }

        .ExternalClass * {
            line-height: 100%;
        }

        body {
            margin: 0;
            padding: 0;
            -webkit-text-size-adjust: 100%;
            -ms-text-size-adjust: 100%;
        }

        table,
        td {
            border-collapse: collapse;
            mso-table-lspace: 0pt;
            mso-table-rspace: 0pt;
        }

        img {
            border: 0;
            height: auto;
            line-height: 100%;
            outline: none;
            text-decoration: none;
            -ms-interpolation-mode: bicubic;
        }

        p {
            display: block;
            margin: 13px 0;
        }
    </style>
    <!--[if !mso]><!-->
    <style type="text/css">
        @media only screen and (max-width:480px) {
            @-ms-viewport {
                width: 320px;
            }

            @viewport {
                width: 320px;
            }
        }
    </style>
    <!--<![endif]-->
    <!--[if mso]><xml>  <o:OfficeDocumentSettings>    <o:AllowPNG/>    <o:PixelsPerInch>96</o:PixelsPerInch>  </o:OfficeDocumentSettings></xml><![endif]-->
    <!--[if lte mso 11]><style type="text/css">  .outlook-group-fix {    width:100% !important;  }</style><![endif]-->
    <!--[if !mso]><!-->
    <link href="https://fonts.googleapis.com/css?family=Ubuntu:300,400,500,700" rel="stylesheet" type="text/css">
    <style type="text/css">
        @import url(https://fonts.googleapis.com/css?family=Ubuntu:300,400,500,700);
    </style>
    <!--<![endif]-->
    <style type="text/css">
        @media only screen and (min-width:480px) {
            .mj-column-per-100 {
                width: 100% !important;
            }
        }
    </style>
</head>
"""
BODY="""\
<body style="background: #FFFFFF;">
    <div class="mj-container" style="background-color:#FFFFFF;">
        <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]-->
        <div style="margin:0px auto;max-width:600px;background:#FFFFFF;">
            <table role="presentation" cellpadding="0" cellspacing="0"
                style="font-size:0px;width:100%;background:#FFFFFF;" align="center" border="0">
                <tbody>
                    <tr>
                        <td
                            style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;">
                            <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]-->
                            <div class="mj-column-per-100 outlook-group-fix"
                                style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;">
                                <table role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
                                    <tbody>
                                        <tr>
                                            <td style="word-wrap:break-word;font-size:0px;padding:0px 0px 0px 0px;"
                                                align="center">
                                                <table role="presentation" cellpadding="0" cellspacing="0"
                                                    style="border-collapse:collapse;border-spacing:0px;" align="center"
                                                    border="0">
                                                    <tbody>
                                                        <tr>
                                                            <td style="width:600px;"><img alt height="auto"
                                                                    src="cid:image1"
                                                                    style="border:none;border-radius:0px;display:block;font-size:13px;outline:none;text-decoration:none;width:100%;height:auto;"
                                                                    width="600"></td>
                                                        </tr>
                                                    </tbody>
                                                </table>
                                            </td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                            <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
        <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]-->
        <div style="margin:0px auto;max-width:600px;background:#FFFFFF;">
            <table role="presentation" cellpadding="0" cellspacing="0"
                style="font-size:0px;width:100%;background:#FFFFFF;" align="center" border="0">
                <tbody>
                    <tr>
                        <td
                            style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;">
                            <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]-->
                            <div class="mj-column-per-100 outlook-group-fix"
                                style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;">
                                <table role="presentation" cellpadding="0" cellspacing="0" style="vertical-align:top;"
                                    width="100%" border="0">
                                    <tbody>
                                        <tr>
                                            <td
                                                style="word-wrap:break-word;font-size:0px;padding:10px 25px;padding-top:10px;padding-right:10px;">
                                                <p
                                                    style="font-size:1px;margin:0px auto;border-top:1px solid #000;width:100%;">
                                                </p>
                                                <!--[if mso | IE]><table role="presentation" align="center" border="0" cellpadding="0" cellspacing="0" style="font-size:1px;margin:0px auto;border-top:1px solid #000;width:100%;" width="600"><tr><td style="height:0;line-height:0;"> </td></tr></table><![endif]-->
                                            </td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                            <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
        <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]-->
        <div style="margin:0px auto;max-width:600px;background:#FFFFFF;">
            <table role="presentation" cellpadding="0" cellspacing="0"
                style="font-size:0px;width:100%;background:#FFFFFF;" align="center" border="0">
                <tbody>
                    <tr>
                        <td
                            style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;">
                            <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]-->
                            <div class="mj-column-per-100 outlook-group-fix"
                                style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;">
                                <table role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
                                    <tbody>
                                        <tr>
                                            <td style="word-wrap:break-word;font-size:0px;padding:15px 15px 15px 15px;"
                                                align="center">
                                                <div
                                                    style="cursor:auto;color:#000000;font-family:Ubuntu, Helvetica, Arial, sans-serif;font-size:11px;line-height:1.5;text-align:center;">
                                                    <p><span style="font-size:28px;"><strong>PMI TOUR CONO SUR 2019 -
                                                                ROSARIO</strong></span></p>
                                                </div>
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="word-wrap:break-word;font-size:0px;padding:15px 15px 15px 15px;"
                                                align="justify">
                                                <div
                                                    style="cursor:auto;color:#000000;font-family:Ubuntu, Helvetica, Arial, sans-serif;font-size:11px;line-height:1.5;text-align:justify;">
                                                    <p>Hola <strong>{nombre}</strong>,</p>
                                                    <p>PMI cumple 50 a&#xF1;os y lo festeja&#xA0; el <strong>5 de
                                                            Noviembre&#xA0;</strong>&#xA0;con conferencistas
                                                        internacionales y de primer nivel.</p>
                                                    <p>Como ya hemos tenido el placer de contar con tu precensia y
                                                        adem&#xE1;s no queremos que te pierdas el Tour de este a&#xF1;o,
                                                        te ofrecemos un c&#xF3;digo promocional especial de
                                                        <strong>$350</strong> (precio actual de <strong>$650</strong>)
                                                        s&#xF3;lo para vos.</p>
                                                    <p>Si te interesa, te pedimos que nos escribas para poder
                                                        facilitarte el c&#xF3;digo de promoci&#xF3;n personalizado para
                                                        que lo ingreses al momento de comprar tu entrada.</p>
                                                    <p>Esperamos poder contar con vos nuevamente y que nos
                                                        acompa&#xF1;es en el festejo de los 50 a&#xF1;itos.</p>
                                                    <p>Saludos y que tengas un bien d&#xED;a.</p>
                                                </div>
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="word-wrap:break-word;font-size:0px;padding:15px 15px 15px 15px;"
                                                align="right">
                                                <div
                                                    style="cursor:auto;color:#000000;font-family:Ubuntu, Helvetica, Arial, sans-serif;font-size:11px;line-height:1.5;text-align:right;">
                                                    <p>Equipo de Voluntarios PMIBA - Comunidad Rosario</p>
                                                    <p><a href="http://www.pmi.org.ar">PMI - Capitulo Buenos Aires -
                                                            Argentina</a></p>
                                                    <p><a href="http://www.pmi.org.ar">www.pmi.org.ar</a></p>
                                                </div>
                                            </td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                            <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
        <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]-->
        <div style="margin:0px auto;max-width:600px;background:#FFFFFF;">
            <table role="presentation" cellpadding="0" cellspacing="0"
                style="font-size:0px;width:100%;background:#FFFFFF;" align="center" border="0">
                <tbody>
                    <tr>
                        <td
                            style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;">
                            <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]-->
                            <div class="mj-column-per-100 outlook-group-fix"
                                style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;">
                                <table role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
                                    <tbody>
                                        <tr>
                                            <td
                                                style="word-wrap:break-word;font-size:0px;padding:10px 25px;padding-top:10px;padding-right:10px;">
                                                <p
                                                    style="font-size:1px;margin:0px auto;border-top:1px solid #000;width:100%;">
                                                </p>
                                                <!--[if mso | IE]><table role="presentation" align="center" border="0" cellpadding="0" cellspacing="0" style="font-size:1px;margin:0px auto;border-top:1px solid #000;width:100%;" width="600"><tr><td style="height:0;line-height:0;"> </td></tr></table><![endif]-->
                                            </td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                            <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
        <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]-->
        <div style="margin:0px auto;max-width:600px;">
            <table role="presentation" cellpadding="0" cellspacing="0" style="font-size:0px;width:100%;" align="center"
                border="0">
                <tbody>
                    <tr>
                        <td
                            style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;">
                            <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]-->
                            <div class="mj-column-per-100 outlook-group-fix"
                                style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;">
                                <table role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
                                    <tbody>
                                        <tr>
                                            <td style="word-wrap:break-word;font-size:0px;padding:0px 0px 0px 0px;"
                                                align="center">
                                                <table role="presentation" cellpadding="0" cellspacing="0"
                                                    style="border-collapse:collapse;border-spacing:0px;" align="center"
                                                    border="0">
                                                    <tbody>
                                                        <tr>
                                                            <td style="width:600px;"><img alt height="auto"
                                                                    src="cid:image2"
                                                                    style="border:none;border-radius:0px;display:block;font-size:13px;outline:none;text-decoration:none;width:100%;height:auto;"
                                                                    width="600"></td>
                                                        </tr>
                                                    </tbody>
                                                </table>
                                            </td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                            <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
        <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]-->
        <div style="margin:0px auto;max-width:600px;">
            <table role="presentation" cellpadding="0" cellspacing="0" style="font-size:0px;width:100%;" align="center"
                border="0">
                <tbody>
                    <tr>
                        <td
                            style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;">
                            <!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]-->
                            <div class="mj-column-per-100 outlook-group-fix"
                                style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;">
                                <table role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
                                    <tbody>
                                        <tr>
                                            <td style="word-wrap:break-word;font-size:0px;padding:0px 0px 0px 0px;"
                                                align="center">
                                                <table role="presentation" cellpadding="0" cellspacing="0"
                                                    style="border-collapse:collapse;border-spacing:0px;" align="center"
                                                    border="0">
                                                    <tbody>
                                                        <tr>
                                                            <td style="width:600px;"><img alt height="auto"
                                                                    src="cid:image3"
                                                                    style="border:none;border-radius:0px;display:block;font-size:13px;outline:none;text-decoration:none;width:100%;height:auto;"
                                                                    width="600"></td>
                                                        </tr>
                                                    </tbody>
                                                </table>
                                            </td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                            <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!--[if mso | IE]>      </td></tr></table>      <![endif]-->
    </div>
</body>

</html>
"""
#----------------------------------------A V I S O -------------------------------------------------------
print('\n')
print("-----------------------------¡ATENCION!--------------------------------")
print("Para utilizar este scrip el arhivo de excel debe estar en la misma ubicación que el script,")
print("además el archivo no debe tener celdas con Nombre y Apellido separados y no contener")
print("campos vacios")
print("")
print("Ingrese el nombre del archivo: ")
archivo= input().format()
print("Ingrese la letra de la columba de Nombres: ")
columnaNombres= input().upper()
print("Ingrese la letra de la columba de Apellidos: ")
columnaApellidos= input().upper()
print("Ingrese la letra de la columba de Mails: ")
columnaMails= input().upper()
#---------------------------------------------------------------------------------------------------------

class asistente:
    def __init__(self, nombre=None, apellido=None, mail=None, type=None):
        self.nombre=nombre
        self.apellido=apellido
        self.mail=mail
        self.type=None

# Abro el excel y obtengo la primer Hoja
book = excel.load_workbook('test2.xlsx')
firstsheet = book.sheetnames[0]
worksheet = book.get_sheet_by_name(firstsheet)

#Defino los arreglos donde voy a extraer cada columna
nombres = []
apellidos = []
mails = []

try:
   #Itero sobre las filas de una determinada columna
    for row in range(2,worksheet.max_row+1):  
        for column in columnaApellidos:  #Recupero los datos de la columna de Apellidos
            cell_name = "{}{}".format(column, row)
            apellidos.append(worksheet[cell_name].value) 

        for column in columnaNombres:  #Recupero los datos de la columna de Nombres
            cell_name = "{}{}".format(column, row)
            nombres.append(worksheet[cell_name].value) 

        for column in columnaMails:  #Recupero los datos de la columna de Mails
            cell_name = "{}{}".format(column, row)
            mails.append(worksheet[cell_name].value) 
        
    Asistentes = []
    #Creo un arreglo con cada asistente dentro
    for i in range (0,worksheet.max_row-1):
        Asistentes.append(asistente(nombres[i],apellidos[i],mails[i]))

    #----------------- M A I L   P E R S O N A L I Z A D O -----------------------------------------------
    #Prototipo de mail (ACA SE PERSONALIZA EL MAIL)
    for i in range (0,worksheet.max_row-1):

        message = MIMEMultipart("alternative")
        message["Subject"] = "PMI TOUR CONO SUR ROSARIO 2019"
        message["From"] = user
        message["To"] = Asistentes[i].mail

        #print(message)
        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(HEAD, "html")
        part2 = MIMEText(BODY.format(nombre=Asistentes[i].nombre), "html") #Acá cambio el nombre del HTML por el personalizado

        # Add HTML/plain-text parts to MIMEMultipart message
        # The email client will try to render the last part first
        message.attach(part1)       
        message.attach(part2)
        
        #Ahora agrego las imágenes que se usan en el HTML
        #This example assumes the image is in the current directory
        fp = open('image1.jpg', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image1>')
        message.attach(msgImage)

        fp = open('image2.jpg', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image2>')
        message.attach(msgImage)

        fp = open('image3.jpg', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image3>')
        message.attach(msgImage)

        
        # Create a secure SSL context
        context = ssl.create_default_context()

        # Try to log in to server and send email
        try:
            server = SMTP(smtp_server,port)
            server.ehlo() # Can be omitted
            #server.starttls(context=context) # Secure the connection
            server.ehlo() # Can be omitted
            server.login(user, password)
            # TODO: Send email here
            #print(message)
            server.sendmail(user, Asistentes[i].mail, message.as_string())

            print(chr(27)+"[1;32m"+"\nMails enviados exitosamente:",(i+1))
            print(chr(27)+"[1;34m"+"Ultimo mail enviado a: " + Asistentes[i].nombre + " " + Asistentes[i].apellido)
            print(chr(27)+"[1;34m"+"mail: " + Asistentes[i].mail+'\033[0;m')
        except Exception as e:
            # Print any error messages to stdout
            print(e)
            print('\033[1;31m'+"\nERROR"+'\033[0;m')
            print(chr(27)+"[1;31m"+"Mails enviados exitosamente:"+'\033[0;m',(chr(27)+"[1;32m"+(i+1)))
            print(chr(27)+"[1;31m"+"Ultimo mail enviado a: " + Asistentes[i].nombre + " " + Asistentes[i].apellido)
            print(chr(27)+"[1;31m"+"mail: " + Asistentes[i].mail+'\033[0;m')
            
        finally:
            server.quit()
            time.sleep(60)# Tiempo [seg] de espera entre mail y mail para evitar entrar en lista negra de SPAM
            #message="" 
        
    #-----------------------------------------------------------------------------------------------------
except (IndexError, TypeError):
    print ("\n ERROR: Verifique lo siguiente:")
    print("+ Que los nombres de las columnas estén bien ingresados")
    print("+ Que en ninguna columna haya un valor vacío")

except:
    print("\n ERROR: Verifique el nombre del archivo")