import openpyxl as excel
import os
from smtplib import SMTP
from email.headerregistry import Address
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import time
# ---------------------------------- D E C L A R A C I O N E S ----------------------------------------
# Para el servidor de correo
port = int(input("Enter your server port (25 by default): ") or "25")


smtp_server = input("Enter your webmail url: ")


user = input("Enter user:")
password = input("Enter password:")
delayBetweenMails = int(input("Enter the delay in seconds netween mails (60 by default): ") or "60")

#message = MIMEMultipart("alternative")
#message["Subject"] = "Test mail personalizado"
#message["From"] = sender_email
#message["To"] = receiver_email

# If you want to send a mail with HTML format your have to separate the HEAD and de BODY
# Separar el HEAD y el BODY para poder editar tranquilamente el BODY con la funcion format()
HEAD = """\
<!DOCTYPE html
	PUBLIC "-//W3C//DTD XHTML 1.0 Transitional //EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">

<html xmlns="http://www.w3.org/1999/xhtml" xmlns:o="urn:schemas-microsoft-com:office:office"
	xmlns:v="urn:schemas-microsoft-com:vml">

<head>
	<!--[if gte mso 9]><xml><o:OfficeDocumentSettings><o:AllowPNG/><o:PixelsPerInch>96</o:PixelsPerInch></o:OfficeDocumentSettings></xml><![endif]-->
	<meta content="text/html; charset=utf-8" http-equiv="Content-Type" />
	<meta content="width=device-width" name="viewport" />
	<!--[if !mso]><!-->
	<meta content="IE=edge" http-equiv="X-UA-Compatible" />
	<!--<![endif]-->
	<title></title>
	<!--[if !mso]><!-->
	<link href="https://fonts.googleapis.com/css?family=Bitter" rel="stylesheet" type="text/css" />
	<link href="https://fonts.googleapis.com/css?family=Open+Sans" rel="stylesheet" type="text/css" />
	<!--<![endif]-->
	<style type="text/css">
		body {
			margin: 0;
			padding: 0;
		}

		table,
		td,
		tr {
			vertical-align: top;
			border-collapse: collapse;
		}

		* {
			line-height: inherit;
		}

		a[x-apple-data-detectors=true] {
			color: inherit !important;
			text-decoration: none !important;
		}
	</style>
	<style id="media-query" type="text/css">
		@media (max-width: 625px) {

			.block-grid,
			.col {
				min-width: 320px !important;
				max-width: 100% !important;
				display: block !important;
			}

			.block-grid {
				width: 100% !important;
			}

			.col {
				width: 100% !important;
			}

			.col>div {
				margin: 0 auto;
			}

			img.fullwidth,
			img.fullwidthOnMobile {
				max-width: 100% !important;
			}

			.no-stack .col {
				min-width: 0 !important;
				display: table-cell !important;
			}

			.no-stack.two-up .col {
				width: 50% !important;
			}

			.no-stack .col.num4 {
				width: 33% !important;
			}

			.no-stack .col.num8 {
				width: 66% !important;
			}

			.no-stack .col.num4 {
				width: 33% !important;
			}

			.no-stack .col.num3 {
				width: 25% !important;
			}

			.no-stack .col.num6 {
				width: 50% !important;
			}

			.no-stack .col.num9 {
				width: 75% !important;
			}

			.video-block {
				max-width: none !important;
			}

			.mobile_hide {
				min-height: 0px;
				max-height: 0px;
				max-width: 0px;
				display: none;
				overflow: hidden;
				font-size: 0px;
			}

			.desktop_hide {
				display: block !important;
				max-height: none !important;
			}
		}
	</style>
</head>
"""
BODY = """\
<body class="clean-body" style="margin: 0; padding: 0; -webkit-text-size-adjust: 100%; background-color: #FFFFFF;">
	<!--[if IE]><div class="ie-browser"><![endif]-->
	<table bgcolor="#FFFFFF" cellpadding="0" cellspacing="0" class="nl-container" role="presentation"
		style="table-layout: fixed; vertical-align: top; min-width: 320px; Margin: 0 auto; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #FFFFFF; width: 100%;"
		valign="top" width="100%">
		<tbody>
			<tr style="vertical-align: top;" valign="top">
				<td style="word-break: break-word; vertical-align: top;" valign="top">
					<!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td align="center" style="background-color:#FFFFFF"><![endif]-->
					<div style="background-color:#FFFFFF;">
						<div class="block-grid mixed-two-up"
							style="Margin: 0 auto; min-width: 320px; max-width: 605px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
							<div
								style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
								<!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#FFFFFF;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:605px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
								<!--[if (mso)|(IE)]><td align="center" width="453" style="background-color:transparent;width:453px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:5px; padding-bottom:5px;"><![endif]-->
								<div class="col num9"
									style="display: table-cell; vertical-align: top; min-width: 320px; max-width: 450px; width: 453px;">
									<div style="width:100% !important;">
										<!--[if (!mso)&(!IE)]><!-->
										<div
											style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:5px; padding-right: 0px; padding-left: 0px;">
											<!--<![endif]-->
											<div align="center" class="img-container center fixedwidth"
												style="padding-right: 0px;padding-left: 0px;">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="center"><![endif]--><img
													align="center" alt="Image" border="0" class="center fixedwidth"
													src="cid:image1"
													style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 431px; display: block;"
													title="Image" width="431" />
												<!--[if mso]></td></tr></table><![endif]-->
											</div>
											<!--[if (!mso)&(!IE)]><!-->
										</div>
										<!--<![endif]-->
									</div>
								</div>
								<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
								<!--[if (mso)|(IE)]></td><td align="center" width="151" style="background-color:transparent;width:151px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:5px; padding-bottom:5px;"><![endif]-->
								<div class="col num3"
									style="display: table-cell; vertical-align: top; max-width: 320px; min-width: 150px; width: 151px;">
									<div style="width:100% !important;">
										<!--[if (!mso)&(!IE)]><!-->
										<div
											style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:5px; padding-right: 0px; padding-left: 0px;">
											<!--<![endif]-->
											<div align="center"
												class="img-container center fullwidthOnMobile fixedwidth"
												style="padding-right: 0px;padding-left: 0px;">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="center"><![endif]--><img
													align="center" alt="Image" border="0"
													class="center fullwidthOnMobile fixedwidth"
													src="cid:image2"
													style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 151px; display: block;"
													title="Image" width="151" />
												<!--[if mso]></td></tr></table><![endif]-->
											</div>
											<!--[if (!mso)&(!IE)]><!-->
										</div>
										<!--<![endif]-->
									</div>
								</div>
								<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
								<!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
							</div>
						</div>
					</div>
					<div style="background-color:#F3F3F3;">
						<div class="block-grid"
							style="Margin: 0 auto; min-width: 320px; max-width: 605px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
							<div
								style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
								<!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#F3F3F3;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:605px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
								<!--[if (mso)|(IE)]><td align="center" width="605" style="background-color:transparent;width:605px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:0px; padding-bottom:0px;"><![endif]-->
								<div class="col num12"
									style="min-width: 320px; max-width: 605px; display: table-cell; vertical-align: top; width: 605px;">
									<div style="width:100% !important;">
										<!--[if (!mso)&(!IE)]><!-->
										<div
											style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:0px; padding-bottom:0px; padding-right: 0px; padding-left: 0px;">
											<!--<![endif]-->
											<table border="0" cellpadding="0" cellspacing="0" class="divider"
												role="presentation"
												style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
												valign="top" width="100%">
												<tbody>
													<tr style="vertical-align: top;" valign="top">
														<td class="divider_inner"
															style="word-break: break-word; vertical-align: top; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%; padding-top: 10px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px;"
															valign="top">
															<table align="center" border="0" cellpadding="0"
																cellspacing="0" class="divider_content" height="0"
																role="presentation"
																style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; border-top: 1px solid #BBBBBB; height: 0px; width: 100%;"
																valign="top" width="100%">
																<tbody>
																	<tr style="vertical-align: top;" valign="top">
																		<td height="0"
																			style="word-break: break-word; vertical-align: top; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
																			valign="top"><span></span></td>
																	</tr>
																</tbody>
															</table>
														</td>
													</tr>
												</tbody>
											</table>
											<div align="center" class="img-container center autowidth fullwidth"
												style="padding-right: 0px;padding-left: 0px;">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="center"><![endif]--><img
													align="center" alt="Image" border="0"
													class="center autowidth fullwidth"
													src="https://d15k2d11r6t6rl.cloudfront.net/public/users/BeeFree/beefree-hj3ygbuwbyn/Flyer%20para%20Mail.png"
													style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 605px; display: block;"
													title="Image" width="605" />
												<!--[if mso]></td></tr></table><![endif]-->
											</div>
											<!--[if (!mso)&(!IE)]><!-->
										</div>
										<!--<![endif]-->
									</div>
								</div>
								<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
								<!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
							</div>
						</div>
					</div>
					<div style="background-color:#F3F3F3;">
						<div class="block-grid"
							style="Margin: 0 auto; min-width: 320px; max-width: 605px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
							<div
								style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
								<!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#F3F3F3;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:605px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
								<!--[if (mso)|(IE)]><td align="center" width="605" style="background-color:transparent;width:605px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:5px; padding-bottom:5px;"><![endif]-->
								<div class="col num12"
									style="min-width: 320px; max-width: 605px; display: table-cell; vertical-align: top; width: 605px;">
									<div style="width:100% !important;">
										<!--[if (!mso)&(!IE)]><!-->
										<div
											style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:5px; padding-right: 0px; padding-left: 0px;">
											<!--<![endif]-->
											<table border="0" cellpadding="0" cellspacing="0" class="divider"
												role="presentation"
												style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
												valign="top" width="100%">
												<tbody>
													<tr style="vertical-align: top;" valign="top">
														<td class="divider_inner"
															style="word-break: break-word; vertical-align: top; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%; padding-top: 10px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px;"
															valign="top">
															<table align="center" border="0" cellpadding="0"
																cellspacing="0" class="divider_content"
																role="presentation"
																style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; border-top: 1px solid #BBBBBB; width: 100%;"
																valign="top" width="100%">
																<tbody>
																	<tr style="vertical-align: top;" valign="top">
																		<td style="word-break: break-word; vertical-align: top; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
																			valign="top"><span></span></td>
																	</tr>
																</tbody>
															</table>
														</td>
													</tr>
												</tbody>
											</table>
											<div align="center" class="img-container center autowidth"
												style="padding-right: 0px;padding-left: 0px;">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="center"><![endif]--><img
													align="center" alt="Image" border="0" class="center autowidth"
													src="cid:image3"
													style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 400px; display: block;"
													title="Image" width="400" />
												<!--[if mso]></td></tr></table><![endif]-->
											</div>
											<!--[if (!mso)&(!IE)]><!-->
										</div>
										<!--<![endif]-->
									</div>
								</div>
								<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
								<!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
							</div>
						</div>
					</div>
					<div style="background-color:#F3F3F3;">
						<div class="block-grid"
							style="Margin: 0 auto; min-width: 320px; max-width: 605px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
							<div
								style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
								<!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#F3F3F3;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:605px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
								<!--[if (mso)|(IE)]><td align="center" width="605" style="background-color:transparent;width:605px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:5px; padding-bottom:5px;"><![endif]-->
								<div class="col num12"
									style="min-width: 320px; max-width: 605px; display: table-cell; vertical-align: top; width: 605px;">
									<div style="width:100% !important;">
										<!--[if (!mso)&(!IE)]><!-->
										<div
											style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:5px; padding-right: 0px; padding-left: 0px;">
											<!--<![endif]-->
											<table border="0" cellpadding="0" cellspacing="0" class="divider"
												role="presentation"
												style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
												valign="top" width="100%">
												<tbody>
													<tr style="vertical-align: top;" valign="top">
														<td class="divider_inner"
															style="word-break: break-word; vertical-align: top; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%; padding-top: 10px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px;"
															valign="top">
															<table align="center" border="0" cellpadding="0"
																cellspacing="0" class="divider_content"
																role="presentation"
																style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; border-top: 1px solid #BBBBBB; width: 100%;"
																valign="top" width="100%">
																<tbody>
																	<tr style="vertical-align: top;" valign="top">
																		<td style="word-break: break-word; vertical-align: top; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
																			valign="top"><span></span></td>
																	</tr>
																</tbody>
															</table>
														</td>
													</tr>
												</tbody>
											</table>
											<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 10px; padding-left: 10px; padding-top: 10px; padding-bottom: 10px; font-family: Georgia, 'Times New Roman', serif"><![endif]-->
											<div
												style="color:#134C75;font-family:'Bitter', Georgia, Times, 'Times New Roman', serif;line-height:1.2;padding-top:10px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
												<div
													style="font-size: 12px; line-height: 1.2; font-family: 'Bitter', Georgia, Times, 'Times New Roman', serif; color: #134C75; mso-line-height-alt: 14px;">
													<p
														style="font-size: 14px; line-height: 1.2; mso-line-height-alt: 17px; margin: 0;">
														<strong><span style="font-size: 28px;">Hola
																{nombre}</span></strong></p>
												</div>
											</div>
											<!--[if mso]></td></tr></table><![endif]-->
											<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 10px; padding-left: 10px; padding-top: 10px; padding-bottom: 10px; font-family: Arial, sans-serif"><![endif]-->
											<div
												style="color:#555555;font-family:'Open Sans', Helvetica, Arial, sans-serif;line-height:1.5;padding-top:10px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
												<div
													style="font-size: 12px; line-height: 1.5; font-family: 'Open Sans', Helvetica, Arial, sans-serif; color: #555555; mso-line-height-alt: 18px;">
													<p
														style="font-size: 14px; line-height: 1.5; text-align: justify; mso-line-height-alt: 21px; margin: 0;">
														PMI cumple 50 años, lo festeja con los mejores conferencistas
														internacionales y como hemos tenido el placer de contar con vos
														como voluntario <strong>queremos regalarte</strong> un descuento
														especial.</p>
													<p
														style="font-size: 14px; line-height: 1.5; text-align: justify; mso-line-height-alt: 21px; margin: 0;">
														Ingresá <strong>TourRosario19ExAsist</strong> como código
														promocional al momento de comprar la entrada y te saldrá
														solamente <strong>$195</strong>.</p>
													<p
														style="font-size: 14px; line-height: 1.5; text-align: justify; mso-line-height-alt: 21px; margin: 0;">
														Esperamos que no te lo pierdas y especialmente poder volver a
														vernos.</p>
													<p
														style="font-size: 14px; line-height: 1.5; mso-line-height-alt: 21px; margin: 0;">
														 </p>
												</div>
											</div>
											<!--[if mso]></td></tr></table><![endif]-->
											<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 10px; padding-left: 10px; padding-top: 10px; padding-bottom: 10px; font-family: Arial, sans-serif"><![endif]-->
											<div
												style="color:#555555;font-family:'Open Sans', Helvetica, Arial, sans-serif;line-height:1.8;padding-top:10px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
												<div
													style="font-size: 12px; line-height: 1.8; font-family: 'Open Sans', Helvetica, Arial, sans-serif; color: #555555; mso-line-height-alt: 22px;">
													<p
														style="font-size: 14px; line-height: 1.8; text-align: right; mso-line-height-alt: 25px; margin: 0;">
														<em>Equipo de Voluntarios PMIBA - Comunidad Rosario</em></p>
													<p
														style="font-size: 14px; line-height: 1.8; text-align: right; mso-line-height-alt: 25px; margin: 0;">
														<a href="https://pmi.org.ar/" rel="noopener"
															style="text-decoration: underline; color: #8F8F8F;"
															target="_blank"><em>PMI -Capítulo Buenos Aires -
																Argentina</em></a></p>
													<p
														style="font-size: 14px; line-height: 1.8; text-align: right; mso-line-height-alt: 25px; margin: 0;">
														 </p>
												</div>
											</div>
											<!--[if mso]></td></tr></table><![endif]-->
											<div align="center" class="button-container"
												style="padding-top:10px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-spacing: 0; border-collapse: collapse; mso-table-lspace:0pt; mso-table-rspace:0pt;"><tr><td style="padding-top: 10px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px" align="center"><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="https://www.eventbrite.com.ar/e/pmi-tour-cono-sur-rosario-2019-tickets-64806809913?aff=ebdssbdestsearch&amp;fbclid=IwAR1nKHTb9PBe_93OAaBDuHJjIeikNG9q1vmJv7nKlWaFiFvqkBO8ORP9kQU" style="height:28.5pt; width:126.75pt; v-text-anchor:middle;" arcsize="64%" stroke="false" fillcolor="#49a6e8"><w:anchorlock/><v:textbox inset="0,0,0,0"><center style="color:#ffffff; font-family:Arial, sans-serif; font-size:14px"><![endif]--><a
													href="https://www.eventbrite.com.ar/e/pmi-tour-cono-sur-rosario-2019-tickets-64806809913?aff=ebdssbdestsearch&amp;fbclid=IwAR1nKHTb9PBe_93OAaBDuHJjIeikNG9q1vmJv7nKlWaFiFvqkBO8ORP9kQU"
													style="-webkit-text-size-adjust: none; text-decoration: none; display: inline-block; color: #ffffff; background-color: #49a6e8; border-radius: 24px; -webkit-border-radius: 24px; -moz-border-radius: 24px; width: auto; width: auto; border-top: 1px solid #49a6e8; border-right: 1px solid #49a6e8; border-bottom: 1px solid #49a6e8; border-left: 1px solid #49a6e8; padding-top: 5px; padding-bottom: 5px; font-family: 'Open Sans', Helvetica, Arial, sans-serif; text-align: center; mso-border-alt: none; word-break: keep-all;"
													target="_blank"><span
														style="padding-left:25px;padding-right:25px;font-size:14px;display:inline-block;">
														<span
															style="font-size: 16px; line-height: 2; mso-line-height-alt: 32px;"><span
																style="font-size: 14px; line-height: 28px;">Comprar
																Entrada</span></span>
													</span></a>
												<!--[if mso]></center></v:textbox></v:roundrect></td></tr></table><![endif]-->
											</div>
											<table border="0" cellpadding="0" cellspacing="0" class="divider"
												role="presentation"
												style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
												valign="top" width="100%">
												<tbody>
													<tr style="vertical-align: top;" valign="top">
														<td class="divider_inner"
															style="word-break: break-word; vertical-align: top; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%; padding-top: 30px; padding-right: 10px; padding-bottom: 30px; padding-left: 10px;"
															valign="top">
															<table align="center" border="0" cellpadding="0"
																cellspacing="0" class="divider_content"
																role="presentation"
																style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; border-top: 1px solid #DDDDDD; width: 100%;"
																valign="top" width="100%">
																<tbody>
																	<tr style="vertical-align: top;" valign="top">
																		<td style="word-break: break-word; vertical-align: top; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;"
																			valign="top"><span></span></td>
																	</tr>
																</tbody>
															</table>
														</td>
													</tr>
												</tbody>
											</table>
											<!--[if (!mso)&(!IE)]><!-->
										</div>
										<!--<![endif]-->
									</div>
								</div>
								<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
								<!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
							</div>
						</div>
					</div>
					<div style="background-color:transparent;">
						<div class="block-grid"
							style="Margin: 0 auto; min-width: 320px; max-width: 605px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
							<div
								style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
								<!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:transparent;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:605px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
								<!--[if (mso)|(IE)]><td align="center" width="605" style="background-color:transparent;width:605px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:5px; padding-bottom:5px;"><![endif]-->
								<div class="col num12"
									style="min-width: 320px; max-width: 605px; display: table-cell; vertical-align: top; width: 605px;">
									<div style="width:100% !important;">
										<!--[if (!mso)&(!IE)]><!-->
										<div
											style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:5px; padding-right: 0px; padding-left: 0px;">
											<!--<![endif]-->
											<div align="center" class="img-container center autowidth fullwidth">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="" align="center"><![endif]--><img
													align="center" alt="Image" border="0"
													class="center autowidth fullwidth"
													src="https://d15k2d11r6t6rl.cloudfront.net/public/users/BeeFree/beefree-hj3ygbuwbyn/Disertantes.png"
													style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 605px; display: block;"
													title="Image" width="605" />
												<!--[if mso]></td></tr></table><![endif]-->
											</div>
											<div align="center" class="img-container center autowidth fullwidth"
												style="padding-right: 0px;padding-left: 0px;">
												<!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="center"><![endif]--><img
													align="center" alt="Image" border="0"
													class="center autowidth fullwidth" src="cid:image4"
													style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 605px; display: block;"
													title="Image" width="605" />
												<!--[if mso]></td></tr></table><![endif]-->
											</div>
											<!--[if (!mso)&(!IE)]><!-->
										</div>
										<!--<![endif]-->
									</div>
								</div>
								<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
								<!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
							</div>
						</div>
					</div>
					<!--[if (mso)|(IE)]></td></tr></table><![endif]-->
				</td>
			</tr>
		</tbody>
	</table>
	<!--[if (IE)]></div><![endif]-->
</body>

</html>
"""
# ----------------------------------------A V I S O -------------------------------------------------------
print('\n')
print("-----------------------------¡WARNING!--------------------------------")
print("The path of the data base in excel must be the same as the path of this script.")
print("The data base must have at least a Name and Mail column with no empty cell")
print("")
print("Enter the name of the file: ")
archivo = input().format()
print("Enter the letter of the Names Column: ")
columnaNombres = input().upper()
print("Enter the letter of the Last Names Column: ")
columnaApellidos = input().upper()
print("Enter the letter of the mails Column: ")
columnaMails = input().upper()
# ---------------------------------------------------------------------------------------------------------


class asistente:
    def __init__(self, name=None, lastName=None, mail=None, type=None):
        self.name = name
        self.lastName = lastName
        self.mail = mail
        self.type = None


# Abro el excel y obtengo la primer Hoja
book = excel.load_workbook(archivo + '.xlsx')
firstsheet = book.sheetnames[0]
worksheet = book.get_sheet_by_name(firstsheet)

# Defino los arreglos donde voy a extraer cada columna
nombres = []
apellidos = []
mails = []
# inicio=input("Indicar número de fila desde donde comenzar a procesar la BDD")
try:
   # Itero sobre las filas de una determinada columna
    for row in range(2, worksheet.max_row+1):
        for column in columnaApellidos:  # Recupero los datos de la columna de Apellidos
            cell_name = "{}{}".format(column, row)
            apellidos.append(worksheet[cell_name].value)

        for column in columnaNombres:  # Recupero los datos de la columna de Nombres
            cell_name = "{}{}".format(column, row)
            nombres.append(worksheet[cell_name].value)

        for column in columnaMails:  # Recupero los datos de la columna de Mails
            cell_name = "{}{}".format(column, row)
            mails.append(worksheet[cell_name].value)

    Asistentes = []
    # Creo un arreglo con cada asistente dentro
    for i in range(0, worksheet.max_row-1):
        Asistentes.append(asistente(nombres[i], apellidos[i], mails[i]))

    # ----------------- M A I L   P E R S O N A L I Z A D O -----------------------------------------------
    # Prototipo de mail (ACA SE PERSONALIZA EL MAIL)
    for i in range(0, worksheet.max_row-1):

        message = MIMEMultipart("alternative")
        message["Subject"] = "PMI TOUR CONO SUR ROSARIO 2019"
        message["From"] = "rosario@pmi.org.ar PMIBA - Comunidad Rosario"
        message["To"] = Asistentes[i].mail

        # print(message)
        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(HEAD, "html")
        # Acá cambio el nombre del HTML por el personalizado
        part2 = MIMEText(BODY.format(nombre=Asistentes[i].name), "html")

        # Add HTML/plain-text parts to MIMEMultipart message
        # The email client will try to render the last part first
        message.attach(part1)
        message.attach(part2)

        # Here you have to add the images that you want to use in the HTML
        # This example assumes the image is in the current directory
        fp = open('image1.jpeg', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image1>') #Here your have to specify the image tag in the HTML
        message.attach(msgImage)

        fp = open('image2.png', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image2>') #Here your have to specify the image tag in the HTML
        message.attach(msgImage)

        fp = open('image3.gif', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image3>') #Here your have to specify the image tag in the HTML
        message.attach(msgImage)

        fp = open('image4.png', 'rb')
        msgImage = MIMEImage(fp.read())
        fp.close()
        # Define the image's ID as referenced above
        msgImage.add_header('Content-ID', '<image4>') #Here your have to specify the image tag in the HTML
        message.attach(msgImage)

        # Create a secure SSL context
        context = ssl.create_default_context()

        # Try to log in to server and send email
        try:
            server = SMTP(smtp_server, port)
            server.ehlo()  # Can be omitted
            # server.starttls(context=context) # Secure the connection
            server.ehlo()  # Can be omitted
            server.login(user, password)
            # TODO: Send email here
            # print(message)
            server.sendmail(user, Asistentes[i].mail, message.as_string())

            print(chr(27)+"[1;32m"+"\nMails successfully sent:", (i+1))
            print(chr(27)+"[1;34m"+"Last name sent to: " +
                  Asistentes[i].name + " " + Asistentes[i].lastName)
            print(chr(27)+"[1;34m"+"mail: " + Asistentes[i].mail+'\033[0;m')
        except Exception as e:
            # Print any error messages to stdout
            print(e)
            print('\033[1;31m'+"\nERROR"+'\033[0;m')
            print(chr(27)+"[1;31m"+"Mails successfully sent:" +
                  '\033[0;m', (chr(27)+"[1;32m"+(i+1)))
            print(chr(27)+"[1;31m"+"Last name sent to: " +
                  Asistentes[i].name + " " + Asistentes[i].lastName)
            print(chr(27)+"[1;31m"+"mail: " + Asistentes[i].mail+'\033[0;m')

        finally:
            server.quit()
            # Tiempo [seg] de espera entre mail y mail para evitar entrar en lista negra de SPAM
            time.sleep(delayBetweenMails)
            # message=""

    # -----------------------------------------------------------------------------------------------------
except (IndexError, TypeError):
    print("\n ERROR: Make sure you have the next items right:")
    print("+ The letters of the columns are right")
    print("+ No column is empty")

except:
    print("\n ERROR: Check the name of the file")
